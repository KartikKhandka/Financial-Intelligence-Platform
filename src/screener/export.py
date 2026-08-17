import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
from pathlib import Path

from loguru import logger
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from src.screener.engine import ScreenerEngine

def export_screener_results(output_path: str | None = None):
    from src import config
    output_path = output_path if output_path else str(config.OUTPUT_DIR / "screener_output.xlsx")
    engine = ScreenerEngine()

    output_dir = config.OUTPUT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  

    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    metric_col_map = {
        "roe_min": "return_on_equity_pct",
        "de_max": "debt_to_equity",
        "fcf_min": "free_cash_flow_cr",
        "revenue_cagr_5yr_min": "revenue_cagr_5yr",
        "pat_cagr_5yr_min": "pat_cagr_5yr",
        "opm_min": "operating_profit_margin_pct",
        "pe_max": "pe_ratio",
        "pb_max": "pb_ratio",
        "dividend_yield_min": "dividend_yield_pct",
        "icr_min": "interest_coverage",
        "market_cap_min": "market_cap_crore",
        "net_profit_min": "net_profit",
        "eps_cagr_min": "eps_cagr_5yr",
        "asset_turnover_min": "asset_turnover",
        "sales_min": "sales",
        "dividend_payout_max": "dividend_payout_ratio_pct",
        "revenue_cagr_3yr_min": "revenue_cagr_3yr",
    }

    for preset_name, preset_filters in engine.presets.items():
        logger.info(f"Exporting preset: {preset_name}")
        df = engine.run_screener(preset_name)

        ws = wb.create_sheet(title=preset_name[:31])  

        if df.empty:
            ws.append(["No companies matched this screener."])
            continue

        kpi_cols = [
            "company_id",
            "company_name",
            "broad_sector",
            "composite_quality_score",
            "return_on_equity_pct",
            "debt_to_equity",
            "free_cash_flow_cr",
            "revenue_cagr_5yr",
            "pat_cagr_5yr",
            "operating_profit_margin_pct",
            "pe_ratio",
            "pb_ratio",
            "dividend_yield_pct",
            "interest_coverage",
            "market_cap_crore",
            "net_profit",
            "sales",
            "eps_cagr_5yr",
            "asset_turnover",
            "dividend_payout_ratio_pct",
        ]

        kpi_cols = [c for c in kpi_cols if c in df.columns]
        export_df = df[kpi_cols]

        for r_idx, row in enumerate(
            dataframe_to_rows(export_df, index=False, header=True), 1
        ):
            ws.append(row)

        headers = [cell.value for cell in ws[1]]

        for filter_key, threshold in preset_filters.items():
            if filter_key in metric_col_map:
                col_name = metric_col_map[filter_key]
                if col_name in headers:
                    col_idx = headers.index(col_name) + 1

                    is_min = filter_key.endswith("_min")

                    for r_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=r_idx, column=col_idx)
                        if cell.value is not None and isinstance(
                            cell.value, (int, float)
                        ):
                            if is_min:
                                if cell.value >= threshold:
                                    cell.fill = green_fill
                                else:
                                    cell.fill = red_fill
                            else:  
                                if cell.value <= threshold:
                                    cell.fill = green_fill
                                else:
                                    cell.fill = red_fill

    wb.save(output_path)
    logger.info(f"Screener output saved to {output_path}")

if __name__ == "__main__":
    export_screener_results()