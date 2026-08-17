import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '../..')))
import sqlite3
from pathlib import Path

import pandas as pd
from loguru import logger
from openpyxl.styles import Font, PatternFill

def generate_peer_comparison_report(db_path: str | None = None, output_path: str | None = None):
    from src import config
    db_path = db_path if db_path else str(config.DB_PATH)
    output_path = output_path if output_path else str(config.OUTPUT_DIR / "peer_comparison.xlsx")
    """Docstring for generate_peer_comparison_report."""
    db_path = Path(db_path)
    if not db_path.exists():
        logger.error(f"DB not found at {db_path}")
        return

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    logger.info("Generating peer comparison excel report...")

    metrics = [
        "net_profit_margin_pct",
        "operating_profit_margin_pct",
        "return_on_equity_pct",
        "debt_to_equity",
        "interest_coverage",
        "asset_turnover",
        "free_cash_flow_cr",
        "capex_cr",
        "earnings_per_share",
        "book_value_per_share",
        "dividend_payout_ratio_pct",
        "roce",
        "roa",
        "revenue_cagr_3yr",
        "revenue_cagr_5yr",
        "pat_cagr_3yr",
        "pat_cagr_5yr",
        "eps_cagr_5yr",
        "cfo_quality_score",
        "composite_quality_score",
    ]

    with sqlite3.connect(db_path) as conn:
        metrics_sql = ", ".join([f"fr.{m}" for m in metrics])
        query = f"""
        SELECT 
            fr.company_id,
            c.company_name,
            p.peer_group_name,
            p.is_benchmark,
            {metrics_sql}
        FROM financial_ratios fr
        JOIN peer_groups p ON fr.company_id = p.company_id
        JOIN companies c ON fr.company_id = c.company_id
        WHERE fr.year = 2024
        """
        df = pd.read_sql_query(query, conn)

    if df.empty:
        logger.warning("No data found for 2024")
        return

    green_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )
    yellow_fill = PatternFill(
        start_color="FFFF00", end_color="FFFF00", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    gold_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for peer_group, group in df.groupby("peer_group_name"):
            group = group.copy()

            for m in metrics:
                pct_col = f"{m}_pct"
                if group[m].isna().all():
                    group[pct_col] = pd.NA
                    continue

                ranks = group[m].rank(pct=True, na_option="keep")
                if m == "debt_to_equity":
                    ranks = 1.0 - ranks
                group[pct_col] = ranks

            cols = ["company_id", "company_name"]
            for m in metrics:
                cols.extend([m, f"{m}_pct"])

            out_df = group[cols].copy()

            summary = {"company_id": "MEDIAN", "company_name": "Peer Group Median"}
            for m in metrics:
                summary[m] = group[m].median()
                summary[f"{m}_pct"] = (
                    pd.NA
                )  

            out_df = pd.concat([out_df, pd.DataFrame([summary])], ignore_index=True)

            sheet_name = peer_group[:31]
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            pct_col_indices = []
            for col_idx, col_name in enumerate(out_df.columns, 1):
                if str(col_name).endswith("_pct"):
                    pct_col_indices.append(col_idx)

            for row_idx, row in enumerate(out_df.itertuples(index=False), 2):
                is_benchmark = False
                if row.company_id != "MEDIAN":
                    original_row = group[group["company_id"] == row.company_id]
                    if (
                        not original_row.empty
                        and original_row.iloc[0]["is_benchmark"] == 1
                    ):
                        is_benchmark = True

                if is_benchmark:
                    for col_idx in range(1, len(out_df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = gold_fill

                if row.company_id != "MEDIAN":
                    for col_idx in pct_col_indices:
                        val = worksheet.cell(row=row_idx, column=col_idx).value
                        if pd.notna(val) and isinstance(val, (int, float)):
                            if val >= 0.75:
                                worksheet.cell(row=row_idx, column=col_idx).fill = (
                                    green_fill
                                )
                            elif val <= 0.25:
                                worksheet.cell(row=row_idx, column=col_idx).fill = (
                                    red_fill
                                )
                            else:
                                worksheet.cell(row=row_idx, column=col_idx).fill = (
                                    yellow_fill
                                )

            summary_row_idx = len(out_df) + 1  
            for col_idx in range(1, len(out_df.columns) + 1):
                cell = worksheet.cell(row=summary_row_idx, column=col_idx)
                cell.font = Font(bold=True)

    logger.info(f"Report generated successfully at {output_path}")

if __name__ == "__main__":
    generate_peer_comparison_report()